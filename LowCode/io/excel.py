#utils/excel.py
# 1. 批量授权高效：支持多用户+多数据批量授权，内置重复校验，避免冗余数据；
# 2. Excel导出实用：保留筛选功能，样式美观（表头染色、边框、列宽自适应），文件名含日期便于归档；
# 3. 接口兼容：筛选参数与日志查询接口一致，降低使用成本，管理员可无缝切换“查询”与“导出”。
# 1.批量撤销灵活：支持三种常见撤销组合，覆盖大部分管理场景，操作高效；
# 2.大文件导出稳定：异步任务避免超时，进度提示提升用户体验，缓存存储临时数据；
# 3.可扩展性强：异步任务支持横向扩展，文件存储可替换为OSS，进度查询可改为WebSocket实时推送。
# Excel导出进度提示（大文件场景，基于异步任务）
# 核心思路：大文件导出（万级 + 日志）耗时久，采用“异步任务 + 进度查询”模式——前端发起导出请求后，后端异步执行，前端通过轮询查询进度，完成后获取下载链接，避免超时。
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
from lowcode.models.models import LowCodeMethodCallLog


def generate_method_log_excel(queryset):
    """
    生成动态方法调用日志Excel文件
    :param queryset: 日志查询集（已筛选的结果）
    :return: Excel文件对象（BytesIO）
    """
    # 1. 将查询集转为DataFrame（方便数据处理）
    log_data = queryset.values(
        "id", "user__username", "model_name", "method_name",
        "params", "result_status", "result_data", "exception_msg",
        "call_time", "time_cost"
    )
    df = pd.DataFrame(list(log_data))

    # 2. 数据格式化
    df.rename(columns={
        "user__username": "调用用户",
        "model_name": "模型名",
        "method_name": "方法名",
        "params": "调用参数",
        "result_status": "调用结果",
        "result_data": "返回数据",
        "exception_msg": "异常信息",
        "call_time": "调用时间",
        "time_cost": "耗时（秒）"
    }, inplace=True)

    # 转换结果状态为中文
    df["调用结果"] = df["调用结果"].map({"success": "成功", "fail": "失败"})

    # 3. 创建Excel工作簿并设置样式
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "动态方法调用日志"

    # 定义样式（表头加粗、居中，边框）
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = openpyxl.styles.PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    center_alignment = Alignment(horizontal="center", vertical="center")

    # 4. 写入表头并设置样式
    headers = df.columns.tolist()
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border

    # 5. 写入数据并设置样式
    for row_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            # 耗时列右对齐
            if headers[col_idx - 1] == "耗时（秒）":
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(horizontal="left")

    # 6. 自动调整列宽
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # 最大列宽限制为50
        ws.column_dimensions[column_letter].width = adjusted_width

    # 7. 保存到BytesIO对象
    from io import BytesIO
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)  # 重置文件指针到开头

    return excel_buffer